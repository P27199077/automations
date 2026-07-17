import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Plan Data Lists
HYROX_PLAN = [
    "Baseline Testing: Run 1.5 km (comfortable pace). Perform 3 rounds of: 10 Goblet Squats, 10 Walking Lunges, 30s Plank.",
    "Strength Foundation: 3 sets: 8 Kettlebell Squats (heavy), 12 Reverse Lunges (bodyweight), 10 Romanian Deadlifts.",
    "Zone 2 Run: 35 minutes of easy jogging at conversational pace. Walk for 1 min if breathing becomes heavy.",
    "Recovery & Mobility: Full rest day. 15 minutes of dynamic leg stretching (quads/hips).",
    "Grip & Upper Body: 3 sets: 10 Push-ups, 12 Dumbbell Rows, 2x50m Farmers Carries (heavy weights).",
    "Compromised Run Intro: Run 1 km (easy). Immediately do: 30 Air Squats + 20 Walking Lunges. Run 1 km (easy).",
    "Active Recovery: 30-minute light walking.",
    "Running Intervals: 5-min warm-up jog. 4 rounds of: 800m run (moderate pace) + 2 min walking rest.",
    "Leg Power & Wall Balls: 3 sets: 10 Heavy Leg Presses (sled push prep), 12 Wall Ball throws, 45s Wall Sit.",
    "Long Aerobic Base: 45 minutes of slow, steady running (Zone 2).",
    "Rest Day: Complete rest day.",
    "Hyrox Work Capacity: 4 rounds: 100m Farmers Carry (heavy), 10 Burpee Broad Jumps, 500m Row or SkiErg.",
    "Compromised Progression: Run 1.5 km. Immediately: 40 Air Squats + 20 Walking Lunges + 10 Burpees. Run 1 km.",
    "Active Recovery: 30-minute light stroll or easy spin cycling.",
    "Aerobic Threshold: Run 3 km at your target race pace (sustainable but challenging).",
    "Heavy Leg Strength: 4 sets: 8 Heavy Leg Presses, 12 Weighted Walking Lunges, 15 Kettlebell Swings.",
    "Conversational Run: 50 minutes of slow, easy jogging.",
    "Rest Day: Complete rest day.",
    "Transition Practice: 3 rounds of: 500m Row/SkiErg, 50m Farmers Carry, 20 Wall Balls, 20 Sandbag Lunges (minimal rest).",
    "Half-Hyrox Simulation: Run 1 km -> 50 Goblet Squats -> Run 1 km -> 50m Farmers Carry -> Run 1 km -> 25 Burpees -> Run 1 km -> 30 Wall Balls.",
    "Deep Recovery: Full rest day. Epsom salt bath.",
    "Easy Recovery Run: 20 minutes of very slow jogging.",
    "Active Mobility: 20 minutes of yoga, stretching, and core planks.",
    "Zone 2 Run: 30 minutes of easy conversational running.",
    "Rest Day: Complete rest day.",
    "Station Quality Review: 2 rounds (light): 10 squats, 10 lunges, 5 burpees. Focus on flawless technique.",
    "30-Day Hyrox Readiness Test: Full-intensity 10 km total moving workout (alternating 1km run and standard bodyweight stations).",
    "Rest Day: Complete rest day.",
    "Tempo Run: 5 km tempo run (first 1km slow, next 3km at target half-marathon pace, last 1km easy).",
    "Leg Strength Maintenance: 3 sets (moderate): 10 Goblet Squats, 10 Lunges, 15 Calf Raises, 45s Plank.",
    "Conversational Base Run: 60 minutes of easy Zone 2 running.",
    "Rest Day: Complete rest day.",
    "Upper Body Maintenance: 3 sets: 10 Push-ups, 12 Rows, 10 Burpee Broad Jumps.",
    "Double-Digit Long Run: 12 km Long Slow Distance (LSD) Run. Keep HR low, walk-drink every 3 km.",
    "Active Recovery: 30-minute easy walk and foam roll.",
    "Speed Intervals: 10-min warm-up. 5 rounds of: 1 km run (moderately fast) + 2 min walking rest.",
    "Core & Lower Body Mobility: Planks, glute bridges, side lunges, and bird-dogs. No heavy lifting.",
    "Semi-Long Run: 8 km easy running.",
    "Rest Day: Complete rest day.",
    "Functional Maintenance: 3 rounds (low weight): 15 Goblet Squats, 15 Wall Balls, 20 lunges.",
    "Peak Long Run: 15 km Long Slow Distance Run. Practice taking energy gels/fuel every 45 mins.",
    "Deep Recovery: Complete rest and stretching.",
    "Aerobic Threshold Run: 6 km run at target half-marathon pace.",
    "Core & Mobility: Hamstring and hip flexor stretches, core work.",
    "Mid-Week Base Run: 10 km easy Zone 2 running.",
    "Rest Day: Complete rest day.",
    "Hyrox Grip Practice: 3 sets: 100m Farmers Carries (heavy), 10 pushups.",
    "Peak Mileage Test: 18 km Long Slow Distance Run. Keep very steady pace, overcome mental fatigue.",
    "Rest & Recovery: Complete rest day.",
    "Easy Shakeout: 5 km very easy run.",
    "Rest Day: Complete rest day.",
    "Light Tempo Run: 4 km run (2km at half-marathon pace).",
    "Rest Day: Complete rest day.",
    "Active Stretching & Walk: 20-minute light walk, dynamic leg swings. Hydrate heavily.",
    "Pre-HM Rest & Carb Load: Rest. Eat clean carbs & drink water with electrolytes.",
    "🏁 HALF MARATHON RUN (21.1 km): Keep pace conservative. Do not sprint. Finish the full distance.",
    "Post-HM Recovery Walk: 20-minute light walking to flush legs. Stretch out calves and IT bands.",
    "Shakeout Run: 3 km slow jog. 3x50m light strides.",
    "Active Mobility: 15 minutes of foam rolling and yoga stretches.",
    "Final Prep & Hydration: Complete rest. Pack your gear. Double up on electrolytes. (Race day tomorrow!)"
]

DSA_PLAN = [
    "Arrays & Dynamic Arrays: Study memory contiguous storage, static vs dynamic arrays, and O(1) time lookups.",
    "Arrays Complexity: Study insert, search, delete time complexities for array bounds.",
    "Arrays Operations: Practice reverse array, find min/max, and basic array rotations in Python/C++.",
    "LeetCode 217 - Contains Duplicate: Implement a solution using linear sorting or dynamic array checks.",
    "LeetCode 217 (Hash Set): Optimize Contains Duplicate using a Hash Set to reach O(N) time complexity.",
    "Strings Foundations: Understand character storage, ASCII transformations, and immutable string bounds.",
    "Strings Operations: Practice reverse string, checks for anagrams, and string tokenization.",
    "LeetCode 242 - Valid Anagram: Implement a frequency sorting solution (O(N log N)).",
    "LeetCode 242 (Hash Map): Optimize Valid Anagram using an array bucket / hash dictionary (O(N) time).",
    "Strings Review: Re-solve Anagram and Duplicate problems under a 15-minute timebox.",
    "Hash Maps theory: Understand bucket tables, collision resolution (chaining vs linear probing), and O(1) searches.",
    "Hash Sets theory: Study set operations, subsets, unions, intersections, and memory allocations.",
    "LeetCode 1 - Two Sum: Solve using the brute-force nested loop approach (O(N^2)).",
    "LeetCode 1 (One-pass Hash Map): Optimize Two Sum using a hash table to store complements in O(N) time.",
    "Hashing Review: Solve contains duplicates & two sum again to cement Hash Map usage.",
    "Two Pointers technique: Learn the logic of dual pointers moving toward the center on sorted lists.",
    "Two Pointers Mechanics: Study pointer increment/decrement boundaries and how to avoid infinite loops.",
    "LeetCode 125 - Valid Palindrome: Write a solution cleaning non-alphanumeric chars and checking match.",
    "LeetCode 125 (Optimized Pointers): Optimize Valid Palindrome using two pointers moving inwards.",
    "Two Pointers Practice: Practice reversing characters in a string in-place with two pointers.",
    "Sliding Window technique: Learn static vs dynamic size windows, tracking sum/length subsets.",
    "Sliding Window calculations: Learn how to slide the window by subtracting left and adding right elements.",
    "LeetCode 121 - Best Time to Buy/Sell Stock: Solve using nested loop checks (O(N^2)).",
    "LeetCode 121 (One-pass Greedy): Optimize Stock Buy/Sell tracking the lowest price and max profit in O(N).",
    "Sliding Window review: Practice building sub-array sum counters and sliding them.",
    "Stack theory: Understand LIFO (Last In First Out), push/pop bounds, and typical recursive equivalents.",
    "Queue theory: Understand FIFO (First In First Out), enqueue/dequeue time complexity, and circular buffers.",
    "LeetCode 20 - Valid Parentheses: Study the stack mapping for bracket configurations.",
    "LeetCode 20 (Stack Solution): Implement Valid Parentheses using a stack and dictionary map in O(N).",
    "Linear Data Structures Review: Re-solve Valid Parentheses, Palindrome, and Two Sum.",
    "Binary Search theory: Learn divide-and-conquer search spaces on sorted arrays.",
    "Binary Search pointers: Study boundary checks, calculation of mid to prevent overflow: left + (right-left)/2.",
    "LeetCode 704 - Binary Search: Implement iterative binary search on a sorted integer array (O(log N)).",
    "LeetCode 704 (Recursive): Implement binary search recursively to understand call stack allocations.",
    "Binary Search practice: Find target insertion index or bounds using binary search logic.",
    "Linked Lists theory: Learn node structures, head/tail pointer addresses, and recursive pointer traversals.",
    "Linked Lists insertion: Practice inserting and deleting nodes at the head, middle, and tail.",
    "LeetCode 206 - Reverse Linked List: Trace pointer changes (prev, curr, next) on paper.",
    "LeetCode 206 (Iteration): Implement iterative reverse linked list by swapping pointer links in-place.",
    "LeetCode 206 (Recursion): Solve reverse linked list using recursion to build call stack familiarity.",
    "Linked Lists Merge logic: Practice merging two sorted lists on paper, tracing node redirection.",
    "LeetCode 21 - Merge Two Sorted Lists: Implement using a dummy node and iterating through list pointers.",
    "LeetCode 21 (Recursive): Merge two sorted lists using recursion for elegant code execution.",
    "Linked Lists review: Re-solve reverse list and merge list in C++ / Python.",
    "Linked Lists cycles: Learn Floyd's cycle detection algorithm (slow and fast pointers).",
    "Trees foundations: Study binary tree traversals (Pre-order, In-order, Post-order) and node structures.",
    "Trees depth theory: Understand depth-first search (DFS) recursion on hierarchical tree paths.",
    "LeetCode 226 - Invert Binary Tree: Trace recursive node inversions (swapping left and right children).",
    "LeetCode 226 (Recursive Solution): Implement recursive invert binary tree (O(N) time, O(H) space).",
    "LeetCode 226 (BFS Iterative): Invert a binary tree iteratively using a queue (Breadth-First Search).",
    "NeetCode 150 Easy Review: Re-solve Contains Duplicate, Valid Anagram, and Two Sum under time pressure.",
    "NeetCode 150 Easy Review: Re-solve Valid Palindrome, Best Time to Buy Stock, and Valid Parentheses.",
    "NeetCode 150 Easy Review: Re-solve Binary Search, Reverse Linked List, and Merge Lists.",
    "NeetCode 150 Easy Review: Re-solve Invert Binary Tree and practice drawing tree recursions.",
    "DSA Mock Test: Simulate an interview. Solve 2 randomly selected Easy tasks within 35 minutes.",
    "Arrays and Hashing Review: Solve a new array-based Easy question from LeetCode.",
    "Two Pointers & Sliding Window Review: Solve a new string-based Easy question.",
    "Answers Verification: Practice trace logs and print debug lines on parenthesis or buffer arrays.",
    "Linked Lists & Trees Review: Practice reversing a sublist or finding tree max depth.",
    "Final DSA Assessment: Review all core concepts and checklist completion state."
]

FULLSTACK_PLAN = [
    "Wireframing & UI Basics: Study user flows and box hierarchies. Draw a paper layout/wireframe of your landing page.",
    "HTML5 Elements: Learn semantic tags. Build a raw style-less contact page using HTML forms and inputs.",
    "CSS3 Foundations: Study selectors, colors, fonts, and the Box Model. Style your contact page to align form elements.",
    "Layout Systems: Master Flexbox and CSS Grid. Create a card grid layout that collapses nicely on mobile screens.",
    "Tailwind CSS: Study utility classes. Rebuild your card grid layout using only Tailwind utility classes.",
    "JavaScript Foundations: Learn variables, data types, and conditionals. Write basic console scripts.",
    "JS Arrays & Objects: Model products and loop over collections to print details to the console.",
    "Advanced JS & DOM: Study ES6 syntax, selection, and modification. Build an interactive clicker game.",
    "Async JS (Fetch): Study Promises and Async/Await. Fetch public quotes/weather API data and render it.",
    "Intro to React JS: Learn components, JSX, and useState. Build a toggle button or counter component.",
    "React Lifecycle & Lists: Study useEffect. Build a search input filtering a dynamic list of API elements.",
    "Python Core Basics: Learn variables, loops, functions, lists, dictionaries, and list comprehensions.",
    "Python OOP & File I/O: Learn class definitions, inheritance, and reading/writing local JSON database files.",
    "Python API & Automation: Write a Python script using requests to fetch and log current cryptocurrency rates.",
    "Node.js & npm: Learn node runtime, package managers, package.json, and install npm packages.",
    "Express.js Servers: Write a basic backend web server responding with JSON on /api/greet.",
    "NoSQL (MongoDB): Study documents. Connect Express server to MongoDB Atlas and insert collections using Mongoose.",
    "SQL Databases: Learn tables, columns, constraints, and execute basic SQLite/MySQL joins.",
    "Express CRUD API: Build a full REST API for a Todo list database connected to MongoDB.",
    "Python NumPy: Learn vector calculations, reshaping, and statistical operations on arrays.",
    "Python Pandas: Read CSV files, clean null data, filter, and summarize DataFrame data.",
    "Connecting MERN: Set up CORS. Fetch backend REST API tasks and render them inside your React state.",
    "Auth & Security: Implement registration and login endpoints using JSON Web Tokens (JWT) and bcrypt.",
    "Python Data Viz: Plot datasets using lines, bars, and scatter plots in Matplotlib and Seaborn.",
    "Gemini/OpenAI SDK: Write a Python script calling the Gemini API to summarize custom user text inputs.",
    "FastAPI/Flask API: Build a light Python API wrapping Gemini model chat responses.",
    "Full Integration: Hook React Front-End to call Express Backend, which queries the Python FastAPI LLM service.",
    "Deployment: Host React on Vercel, Node server on Render, and MongoDB on Atlas.",
    "shadcn/ui Dashboard: Assemble highly polished dashboard UI pages quickly using shadcn component libraries.",
    "Prebuilt Auth (Clerk): Add Clerk social authentication (Google/GitHub login) to lock/unlock React pages.",
    "Git Collaboration: Practice staging, branching, pull requests, and resolving merge conflicts.",
    "Third-party APIs: Integrate transactional notifications or send billing alerts using Resend SDK.",
    "Boilerplate Starter: Run a production starter kit template (Next.js + Tailwind + Clerk + Prisma).",
    "Tailwind Animations: Add slide-overs, drawer menus, and micro-hover cards using Framer Motion.",
    "Speed-run Challenge: Timebox building a full landing page with email signup in under 4 hours.",
    "AI Streaming Chat: Implement chat response letter-by-letter text streaming using Vercel AI SDK.",
    "Vector Databases: Generate text embeddings and store them in Pinecone / Supabase Vector databases.",
    "RAG (Retrieval Gen): Write a script querying LLMs on custom vector knowledge files.",
    "Gemini Function Calling: Register local Python functions for LLM callbacks.",
    "LangChain Chains: Build a multi-step sequential LLM pipeline (Idea -> Translate -> Tweet).",
    "Image/Audio APIs: Generate illustrative assets using DALL-E/ElevenLabs APIs.",
    "Prompt Engineering: Write safe system prompts for an AI health chatbot.",
    "Hackathon Scoping MVP: Pick a hackathon prompt and scope down to 3 core features.",
    "UI Mockups & Boilerplate: Initialize hackathon database schema and nav bar.",
    "Core DB Coding: Write form input controllers to save backend mock records.",
    "Core AI Coding: Hook up streaming prompts inside the server endpoints.",
    "Polish Visual States: Add loader states, skeletons, and error screens.",
    "Live Hackathon Deploy: Set up domains on Vercel and check console logs.",
    "Pitch Preparation: Structure problem, solution, and mechanism statements.",
    "Landing Page Optimization: Design conversion-optimized Hero block.",
    "Charts & Analytics: Display database telemetry on Chart.js/Recharts in React.",
    "Winning Demo Script: Write a 120s script highlighting the product's value.",
    "Record Loom Demo: Film a seamless walkthrough of the hackathon build.",
    "Write README: Compose an awesome repository landing page with visual assets.",
    "Devpost Submission: Create final Devpost submissions drafts and descriptions.",
    "Portfolio Review: Bundle all month 1 & 2 items into your main landing page.",
    "Hackathon Polish: Refine UI and responsive styling details.",
    "Mock Auditing: Test edge cases and inputs validation.",
    "Pitch Deck Slides: Design simple slides for your product pitch.",
    "Final Launch: Prepare public repositories and share links on social channels."
]

# 2. Setup workbook
wb = Workbook()
ws = wb.active
ws.title = "60-Day Life Reset"

# Show grid lines explicitly
ws.views.sheetView[0].showGridLines = True

# Style definitions
title_fill = PatternFill(start_color="0f172a", end_color="0f172a", fill_type="solid")
header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
zebra_fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
white_fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")

font_title = Font(name="Segoe UI", size=16, bold=True, color="ffffff")
font_header = Font(name="Segoe UI", size=10, bold=True, color="ffffff")
font_body = Font(name="Segoe UI", size=9)
font_bold = Font(name="Segoe UI", size=9, bold=True)
font_muted = Font(name="Segoe UI", size=8.5, color="475569")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border_side = Side(border_style="thin", color="e2e8f0")
thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

# 3. Create Title Block
ws.merge_cells("A1:P2")
title_cell = ws["A1"]
title_cell.value = "60-DAY LIFE RESET GAMIFIED CHALLENGE TRACKER"
title_cell.font = font_title
title_cell.fill = title_fill
title_cell.alignment = align_center

# 4. Create Headers Row (Row 4)
headers = [
    "Day",
    "Date",
    "Wake 5AM",
    "Gym Session",
    "Hyrox Done",
    "Hyrox Activity Workout",
    "4L Water",
    "Protein",
    "LeetCode Done",
    "DSA Study Topic & LeetCode Task",
    "Side Project Done",
    "Full-Stack / AI Learning Task",
    "Korean 40m",
    "Make Reels",
    "100 Core",
    "Fast Typing"
]

for col_idx, text in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx)
    cell.value = text
    cell.font = font_header
    cell.fill = header_fill
    cell.alignment = align_center
    cell.border = thin_border

# Set row heights
ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 20
ws.row_dimensions[3].height = 10  # blank space
ws.row_dimensions[4].height = 26  # header

# 5. Populate Data Rows (Row 5 to 64)
start_date = datetime.date(2026, 5, 25)

for i in range(60):
    row_idx = 5 + i
    current_date = start_date + datetime.timedelta(days=i)
    date_str = current_date.strftime("%a, %b %d")
    
    # Alternating row fill
    row_fill = zebra_fill if i % 2 == 1 else white_fill
    
    # Columns mapping:
    # A: Day, B: Date, C: Wake, D: Gym, E: Hyrox Done, F: Hyrox Text,
    # G: Water, H: Protein, I: LC Done, J: LC Text, K: Project Done, L: Project Text,
    # M: Korean, N: Reels, O: Core, P: Typing
    
    # Build cells
    c_day = ws.cell(row=row_idx, column=1, value=f"Day {i+1}")
    c_day.font = font_bold
    c_day.alignment = align_center
    
    c_date = ws.cell(row=row_idx, column=2, value=date_str)
    c_date.font = font_body
    c_date.alignment = align_center
    
    # Checkbox columns (initialized to FALSE so Google Sheets easily renders checkboxes)
    checkbox_cols = [3, 4, 5, 7, 8, 9, 11, 13, 14, 15, 16]
    for col in checkbox_cols:
        c_check = ws.cell(row=row_idx, column=col, value=False)
        c_check.font = font_body
        c_check.alignment = align_center
        
    # Text Plan columns
    c_hyrox = ws.cell(row=row_idx, column=6, value=HYROX_PLAN[i])
    c_hyrox.font = font_body
    c_hyrox.alignment = align_left
    
    c_dsa = ws.cell(row=row_idx, column=10, value=DSA_PLAN[i])
    c_dsa.font = font_body
    c_dsa.alignment = align_left
    
    c_fs = ws.cell(row=row_idx, column=12, value=FULLSTACK_PLAN[i])
    c_fs.font = font_body
    c_fs.alignment = align_left
    
    # Apply borders & fills across all row cells
    for col_idx in range(1, 17):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        cell.fill = row_fill
        
    ws.row_dimensions[row_idx].height = 42  # ample height for wrapped descriptions

# 6. Adjust Column Widths
col_widths = {
    "A": 10,  # Day
    "B": 13,  # Date
    "C": 12,  # Wake
    "D": 12,  # Gym
    "E": 12,  # Hyrox Done
    "F": 38,  # Hyrox Text
    "G": 11,  # Water
    "H": 11,  # Protein
    "I": 12,  # LC Done
    "J": 38,  # DSA Text
    "K": 13,  # Project Done
    "L": 38,  # Project Text
    "M": 12,  # Korean
    "N": 12,  # Reels
    "O": 11,  # Core
    "P": 12   # Typing
}

for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

output_path = "60_Day_Life_Reset_Planner.xlsx"
wb.save(output_path)
print(f"Excel workbook successfully saved to {output_path}")
